#!/usr/bin/env python3
"""
Excel exporter utility for UAT validation reports.

Generates comprehensive Excel workbooks with:
- Multiple sheets per test category
- Conditional formatting (PASS=Green, FAIL=Red)
- Summary dashboard with statistics
- Formula preservation for manual verification
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from typing import Dict, List, Any
from pathlib import Path


class UATExcelExporter:
    """
    Generates comprehensive Excel workbooks for UAT validation.

    Features:
    - Multiple sheets per test category
    - Conditional formatting (Pass=Green, Fail=Red)
    - Formula preservation for manual verification
    - Summary dashboard with statistics
    """

    # Color schemes
    COLOR_PASS = "C6EFCE"  # Light green
    COLOR_FAIL = "FFC7CE"  # Light red
    COLOR_HEADER = "4472C4"  # Blue
    COLOR_WARNING = "FFEB9C"  # Yellow

    def __init__(self, output_path: str, title: str = "MARGIN & POSITION MANAGEMENT UAT"):
        """
        Initialize Excel exporter.

        Args:
            output_path: Full path to output Excel file
            title: Dashboard title (default: "MARGIN & POSITION MANAGEMENT UAT")
        """
        self.output_path = Path(output_path)
        self.title = title
        self.wb = Workbook()
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        self.summary_data = []
        self.test_results = {}

    def add_test_result(
        self,
        sheet_name: str,
        test_data: pd.DataFrame,
        validation_data: pd.DataFrame,
        description: str = ""
    ):
        """
        Add test result to workbook with validation comparison.

        Args:
            sheet_name: Name of test category (max 31 chars for Excel)
            test_data: Input data and intermediate calculations
            validation_data: Expected vs Actual comparison with Status column
            description: Optional description for the test sheet
        """
        # Truncate sheet name if too long
        sheet_name = sheet_name[:31]

        # Create new sheet
        ws = self.wb.create_sheet(title=sheet_name)

        # Add description if provided
        row_offset = 1
        if description:
            ws.merge_cells(f'A1:E1')
            ws['A1'] = description
            ws['A1'].font = Font(bold=True, size=12)
            row_offset = 3

        # Write test data section
        ws.cell(row=row_offset, column=1, value="Test Inputs & Calculations")
        ws.cell(row=row_offset, column=1).font = Font(bold=True, size=11)
        row_offset += 1

        self._write_dataframe(ws, test_data, row_offset)
        row_offset += len(test_data) + 3

        # Write validation section
        ws.cell(row=row_offset, column=1, value="Validation Results")
        ws.cell(row=row_offset, column=1).font = Font(bold=True, size=11)
        row_offset += 1

        self._write_dataframe(ws, validation_data, row_offset)

        # Apply conditional formatting to validation section
        if 'Status' in validation_data.columns:
            status_col_idx = validation_data.columns.get_loc('Status') + 1
            for row_idx in range(row_offset + 1, row_offset + 1 + len(validation_data)):
                cell = ws.cell(row=row_idx, column=status_col_idx)
                if cell.value == 'PASS':
                    cell.fill = PatternFill(start_color=self.COLOR_PASS, end_color=self.COLOR_PASS, fill_type="solid")
                elif cell.value == 'FAIL':
                    cell.fill = PatternFill(start_color=self.COLOR_FAIL, end_color=self.COLOR_FAIL, fill_type="solid")

        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

        # Calculate summary stats
        pass_count = (validation_data['Status'] == 'PASS').sum() if 'Status' in validation_data.columns else 0
        fail_count = (validation_data['Status'] == 'FAIL').sum() if 'Status' in validation_data.columns else 0
        total_count = len(validation_data)

        self.summary_data.append({
            'Test Category': sheet_name,
            'Total Tests': total_count,
            'Passed': pass_count,
            'Failed': fail_count,
            'Pass Rate': f"{(pass_count/total_count*100):.1f}%" if total_count > 0 else "N/A"
        })

        self.test_results[sheet_name] = validation_data

    def _write_dataframe(self, worksheet, df: pd.DataFrame, start_row: int):
        """
        Write pandas DataFrame to worksheet with formatting.

        Args:
            worksheet: openpyxl worksheet object
            df: DataFrame to write
            start_row: Starting row number
        """
        # Write headers
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = worksheet.cell(row=start_row, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write data
        for row_idx, row_data in enumerate(df.itertuples(index=False), start=start_row + 1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)

                # Format numbers
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    if abs(value) < 1 and value != 0:
                        cell.number_format = '0.0000'  # Small decimals
                    elif abs(value) < 100:
                        cell.number_format = '0.00'
                    else:
                        cell.number_format = '#,##0.00'

                # Center alignment
                cell.alignment = Alignment(horizontal='center', vertical='center')

    def _auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths based on content."""
        from openpyxl.cell.cell import MergedCell

        for column in worksheet.columns:
            max_length = 0
            column_letter = None

            for cell in column:
                # Skip merged cells
                if isinstance(cell, MergedCell):
                    continue

                if column_letter is None:
                    column_letter = cell.column_letter

                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            if column_letter:
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                worksheet.column_dimensions[column_letter].width = adjusted_width

    def add_summary_dashboard(self):
        """Generate summary dashboard sheet."""
        if not self.summary_data:
            return

        # Create summary sheet (insert at beginning)
        ws = self.wb.create_sheet(title="Summary Dashboard", index=0)

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = self.title
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25

        # Date
        import datetime
        ws['A2'] = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True)

        # Overall statistics
        summary_df = pd.DataFrame(self.summary_data)
        total_tests = summary_df['Total Tests'].sum()
        total_passed = summary_df['Passed'].sum()
        total_failed = summary_df['Failed'].sum()
        overall_pass_rate = (total_passed / total_tests * 100) if total_tests > 0 else 0

        ws['A4'] = "Overall Statistics"
        ws['A4'].font = Font(bold=True, size=12)

        stats = [
            ['Total Test Cases:', total_tests],
            ['Passed:', total_passed],
            ['Failed:', total_failed],
            ['Pass Rate:', f"{overall_pass_rate:.1f}%"]
        ]

        for idx, (label, value) in enumerate(stats, start=5):
            ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=idx, column=2, value=value)

            # Color code pass rate
            if label == 'Pass Rate:':
                cell = ws.cell(row=idx, column=2)
                if overall_pass_rate == 100:
                    cell.fill = PatternFill(start_color=self.COLOR_PASS, end_color=self.COLOR_PASS, fill_type="solid")
                elif overall_pass_rate < 100:
                    cell.fill = PatternFill(start_color=self.COLOR_FAIL, end_color=self.COLOR_FAIL, fill_type="solid")

        # Detailed results by category
        ws['A10'] = "Results by Test Category"
        ws['A10'].font = Font(bold=True, size=12)

        self._write_dataframe(ws, summary_df, start_row=11)

        # Error statistics (if any failures)
        if total_failed > 0:
            ws['A' + str(14 + len(summary_df))] = "⚠️ FAILURES DETECTED - Review individual test sheets"
            ws['A' + str(14 + len(summary_df))].font = Font(bold=True, color="FF0000", size=11)
        else:
            ws['A' + str(14 + len(summary_df))] = "✅ ALL TESTS PASSED"
            ws['A' + str(14 + len(summary_df))].font = Font(bold=True, color="008000", size=11)

        # Auto-adjust columns
        self._auto_adjust_columns(ws)

        # Set column A width for labels
        ws.column_dimensions['A'].width = 25

    def save(self):
        """Save workbook to disk."""
        # Ensure output directory exists
        self.output_path.parent.mkdir(parents=True, exist_ok=True)

        # Save
        self.wb.save(self.output_path)
        print(f"\n{'='*70}")
        print(f"[SUCCESS] UAT Report saved: {self.output_path}")
        print(f"{'='*70}\n")
